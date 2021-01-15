import urllib, urllib2, zipfile, os, sys, xlsxwriter, openpyxl
from poster.encode import multipart_encode
from poster.streaminghttp import register_openers
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Alignment
reload(sys)
sys.setdefaultencoding('utf-8')


# 参考: https://my.oschina.net/whp/blog/127909
def upload_file(url, file_full_name):
    register_openers()
    datagen, headers = multipart_encode({'file': open(file_full_name, 'rb')})
    request = urllib2.Request(url, datagen, headers)
    ret = urllib2.urlopen(request).read()
    return ret

def download_file(url, file_full_name):
    urllib.urlretrieve(url, file_full_name)

def zip_directory(directory_path, file_full_name):
    zip = zipfile.ZipFile(file_full_name, 'w', zipfile.ZIP_DEFLATED)
    for directory, directory_name, file_names in os.walk(directory_path):
        file_path = directory.replace(directory_path, '')  # 这一句很重要，不replace的话，就从根目录开始复制
        file_path = file_path and file_path + os.sep or '' # 实现文件夹以及包含的所有文件的压缩
        for file_name in file_names:
            zip.write(os.path.join(directory, file_name).decode(encoding = 'GBK'), file_path + file_name)
    zip.close()

def read_from_excel(file_full_name):
    workbook = openpyxl.load_workbook(file_full_name, read_only = True)
    worksheet = workbook[workbook.sheetnames[0]]  # 默认读取第一个sheet
    titles = []
    ret = []
    row_number = 0
    col_number = 0
    for row in worksheet.rows:
        js = {}
        for cell in row:
            if row_number == 0:
                titles.append(cell.value)
            else:
                js[titles[col_number]] = cell.value
                col_number += 1
        if row_number != 0:
            ret.append(js)
        row_number += 1
        col_number = 0
    return ret

# 写入excel时，xlsxwriter比openpyxl快很多，且生成的excel文档要小10%以上
# 使用openpyxl写入excel时，加上write_only=True参数对性能影响很有限
def write_to_excel(data, file_full_name):
    workbook = xlsxwriter.Workbook(file_full_name)
    worksheet = workbook.add_worksheet() # 创建名为'Sheet1'的sheet
    style_title = workbook.add_format({'bold': True, 'align': 'center', 'font_size': 10})
    style_content = workbook.add_format({'font_size': 9})
    row_number = 0
    col_number = 0
    keys = data[0].keys()
    for key in keys:
        worksheet.write(row_number, col_number, str(key) if key is not None else '', style_title)
        col_number += 1
    for js in data:
        row_number += 1
        col_number = 0
        for key in keys:
            worksheet.write(row_number, col_number, str(js[key]) if js[key] is not None else '', style_content)
            col_number += 1
    workbook.close()

def write_to_excel2(data, file_full_name):
    workbook = openpyxl.Workbook()  # 默认有一个名为'Sheet'的sheet
    workbook.create_sheet()  # 创建名为'Sheet1'的sheet
    workbook.remove(workbook[workbook.sheetnames[0]])  # 删除名为'Sheet'的sheet  以上两行可以不要，只是OCD患者一定要sheet从'Sheet1'开始
    worksheet = workbook.active  # 也可以用 workbook[workbook.sheetnames[0]]
    alignment_title = Alignment(horizontal = 'center')
    font_title = Font(bold = True, size = 10)
    font_content = Font(size = 9)
    row_number = 1
    col_number = 1
    keys = data[0].keys()
    for key in keys:
        worksheet.cell(row_number, col_number, str(key) if key is not None else '')
        worksheet.cell(row_number, col_number).font = font_title
        worksheet.cell(row_number, col_number).alignment = alignment_title
        col_number += 1
    for js in data:
        row_number += 1
        col_number = 1
        for key in keys:
            worksheet.cell(row_number, col_number, str(js[key]) if js[key] is not None else '')
            worksheet.cell(row_number, col_number).font = font_content
            col_number += 1
    workbook.save(file_full_name)

def write_to_excel3(data, file_full_name):
    workbook = openpyxl.Workbook(write_only = True)  # 默认无sheet
    workbook.create_sheet()  # 创建名为'Sheet'的sheet
    worksheet = workbook.create_sheet()  # 创建名为'Sheet1'的sheet
    workbook.remove(workbook[workbook.sheetnames[0]])  # 删除名为'Sheet'的sheet  以上两行可以不要，只是OCD患者一定要sheet从'Sheet1'开始
    alignment_title = Alignment(horizontal = 'center')
    font_title = Font(bold = True, size = 10)
    font_content = Font(size = 9)
    keys = data[0].keys()
    row = []
    for key in keys:
        cell = WriteOnlyCell(worksheet, str(key) if key is not None else '')
        cell.font = font_title
        cell.alignment = alignment_title
        row.append(cell)
    worksheet.append(row)
    for js in data:
        row = []
        for key in keys:
            cell = WriteOnlyCell(worksheet, str(js[key]) if js[key] is not None else '')
            cell.font = font_content
            cell.alignment = alignment_title
            row.append(cell)
        worksheet.append(row)
    workbook.save(file_full_name)
