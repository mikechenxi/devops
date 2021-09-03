import urllib.request, zipfile, os, collections, xlsxwriter, openpyxl, requests, pandas, pyexcel, ssl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Alignment
ssl._create_default_https_context = ssl._create_unverified_context


# 用 urllib.request 实在是搞不定... python3 中 open() 函数返回的不是 file 对象了, 而是流...
def upload_file(url, file_full_name):
    file_name = os.path.basename(file_full_name)
    file = open(file_full_name, 'rb')
    files = {'file': (file_name, file, 'multipart/form-data')}
    res = requests.post(url, files = files)
    file.close()
    return res.text


def download_file(url, file_full_name):
    urllib.urlretrieve(url, file_full_name)

    
def zip_directory(directory_path, file_full_name):
    zip = zipfile.ZipFile(file_full_name, 'w', zipfile.ZIP_DEFLATED)
    for directory, directory_name, file_names in os.walk(directory_path):
        file_path = directory.replace(directory_path, '')  # 这一句很重要，不replace的话，就从根目录开始复制
        file_path = file_path and file_path + os.sep or '' # 实现文件夹以及包含的所有文件的压缩
        for file_name in file_names:
            zip.write(os.path.join(directory, file_name), file_path + file_name)
    zip.close()

    
# [[{},{}],[{},{}]]
def read_from_excel(file_full_name):
    workbook = openpyxl.load_workbook(file_full_name, read_only = True)
    ret = []
    for worksheet in workbook.worksheets:
        titles = []
        data = []
        row_number = 0
        col_number = 0
        for row in worksheet.rows:
            js = collections.OrderedDict()
            for cell in row:
                if row_number == 0:
                    titles.append(cell.value)
                else:
                    js[titles[col_number]] = cell.value
                    col_number += 1
            if row_number != 0:
                data.append(js)
            row_number += 1
            col_number = 0
        ret.append(data)
    return ret


# [{},{}]
def read_from_excel2(file_full_name):
    ret = []
    records = pyexcel.iget_records(file_name = file_full_name)
    for record in records:
        ret.append(record)
    pyexcel.free_resources()
    return ret


# [{},{}]
def read_from_excel3(file_full_name):
    ret = []
    rows = pandas.read_excel(file_full_name)
    columns = rows.columns
    for row in rows.itertuples():
        i = 1
        js = collections.OrderedDict()
        for column in columns:
            js[column] = row[i]
            i += 1
        ret.append(js)
    return ret


# 写入excel时，xlsxwriter比openpyxl快很多，且生成的excel文档要小10%以上
# 使用openpyxl写入excel时，加上write_only=True参数对性能影响很有限
# constant_memory: https://xlsxwriter.readthedocs.io/working_with_memory.html
# [[{},{}],[{},{}]]
def write_to_excel(datas, file_full_name, constant_memory = False):
    workbook = xlsxwriter.Workbook(file_full_name, {'constant_memory': constant_memory})
    for data in datas:
        worksheet = workbook.add_worksheet() # 创建名为'Sheet1'的sheet
        style_title = workbook.add_format({'bold': True, 'align': 'center', 'font_size': 10})
        style_content = workbook.add_format({'font_size': 9})
        row_number = 0
        col_number = 0
        keys = data[0].keys()
        for key in keys:
            worksheet.write(row_number, col_number, str(key) if key is not None else '', style_title)
            col_number += 1
        for js in data:
            row_number += 1
            col_number = 0
            for key in keys:
                worksheet.write(row_number, col_number, str(js[key]) if js[key] is not None else '', style_content)
                col_number += 1
    workbook.close()
    log_util.log('file_util.write_to_excel', 'filename:%s' % (file_full_name))


# [[{},{}],[{},{}]]
def write_to_excel2_1(datas, file_full_name):
    workbook = openpyxl.Workbook()  # 默认有一个名为'Sheet'的sheet
    for data in datas:
        worksheet = workbook.create_sheet()  # 创建名为'Sheet1'的sheet
        alignment_title = Alignment(horizontal = 'center')
        font_title = Font(bold = True, size = 10)
        font_content = Font(size = 9)
        keys = data[0].keys()
        row_number = 1
        col_number = 1
        for key in keys:
            worksheet.cell(row_number, col_number, str(key) if key is not None else '')
            worksheet.cell(row_number, col_number).font = font_title
            worksheet.cell(row_number, col_number).alignment = alignment_title
            col_number += 1
        for js in data:
            row_number += 1
            col_number = 1
            for key in keys:
                worksheet.cell(row_number, col_number, str(js[key]) if js[key] is not None else '')
                worksheet.cell(row_number, col_number).font = font_content
                col_number += 1
    workbook.remove(workbook[workbook.sheetnames[0]])  # 删除默认的名为'Sheet'的sheet
    workbook.save(file_full_name)
    log_util.log('file_util.write_to_excel', 'filename:%s' % (file_full_name))


# [[{},{}],[{},{}]]
def write_to_excel2_2(datas, file_full_name):
    workbook = openpyxl.Workbook(write_only = True)  # 默认无sheet sheet名从'Sheet'开始
    workbook.create_sheet()  # 创建名为'Sheet'的sheet
    for data in datas:
        worksheet = workbook.create_sheet()  # 创建名为'Sheet1'的sheet
        alignment_title = Alignment(horizontal = 'center')
        font_title = Font(bold = True, size = 10)
        font_content = Font(size = 9)
        keys = data[0].keys()
        row = []
        for key in keys:
            cell = WriteOnlyCell(worksheet, str(key) if key is not None else '')
            cell.font = font_title
            cell.alignment = alignment_title
            row.append(cell)
        worksheet.append(row)
        for js in data:
            row = []
            for key in keys:
                cell = WriteOnlyCell(worksheet, str(js[key]) if js[key] is not None else '')
                cell.font = font_content
                cell.alignment = alignment_title
                row.append(cell)
            worksheet.append(row)
    workbook.remove(workbook[workbook.sheetnames[0]])  # 删除名为'Sheet'的sheet
    workbook.save(file_full_name)
    log_util.log('file_util.write_to_excel', 'filename:%s' % (file_full_name))

# [{},{}]
def write_to_excel3(data, file_full_name):
    pyexcel.save_as(records = data, dest_file_name = file_full_name)

# [{},{}]
def write_to_excel4(data, file_full_name):
    df = pandas.DataFrame(data)
    df.to_excel(file_full_name, index = False)
